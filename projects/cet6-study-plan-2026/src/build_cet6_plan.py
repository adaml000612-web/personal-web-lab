from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    KeepTogether,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table as PDFTable,
    TableStyle,
)


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

XLSX_PATH = OUTPUT_DIR / "2026年12月CET-6备考计划_初版_v1.xlsx"
PDF_PATH = OUTPUT_DIR / "2026年12月CET-6备考计划_初版_v1.pdf"

START_DATE = date(2026, 9, 7)
PLACEHOLDER_EXAM_DATE = date(2026, 12, 12)
GENERATED_DATE = date(2026, 8, 27)

NEEA_NEWS_URL = "https://cet.neea.edu.cn/html1/category/16093/1124-1.htm"
NEEA_TEST_URL = "https://cet-kw.neea.edu.cn/Home/TestDatePlan.html"
NEEA_CET_URL = "https://cet.neea.edu.cn/xhtml1/folder/16113/1586-1.htm"
NEEA_2026H1_URL = "https://cet.neea.edu.cn/html1/report/2603/2-1.htm"
HOLIDAY_URL = "https://www.gov.cn/zhengce/zhengceku/202511/content_7047091.htm"
SHANGHAI_PLAN_URL = "https://edu.sh.gov.cn/cmsres/28/286d084bda7e46cea1e7d4702b480c86/19d146e3093dac64c68581d795a298e2.pdf"


NAVY = "173B57"
TEAL = "1F7A8C"
TEAL_DARK = "145B68"
SKY = "DCEEF3"
PALE_BLUE = "EDF6F8"
AMBER = "D99B3D"
PALE_AMBER = "FFF3D8"
GREEN = "2E7D5B"
PALE_GREEN = "E6F3EC"
RED = "B84A4A"
PALE_RED = "FBE9E9"
PURPLE = "6C5B7B"
PALE_PURPLE = "EFEAF3"
INK = "243447"
MUTED = "667085"
LINE = "D5DEE5"
WHITE = "FFFFFF"
OFFWHITE = "F7F9FB"
GREY = "E9EDF1"


WEEKDAY_CN = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]


@dataclass
class DayPlan:
    day: date
    week_no: int
    phase: str
    day_type: str
    planned_minutes: int
    word: str = ""
    listening: str = ""
    reading: str = ""
    writing_translation: str = ""
    review: str = ""
    material: str = ""
    success: str = ""
    status: str = "未开始"


WEEK_SPECS = {
    1: ("诊断启动", "建立基线，不追求高分；把错因分成词汇、听音、定位、逻辑、速度五类。", "F0"),
    2: ("基础Ⅰ：长对话 + 仔细阅读", "长对话抓人物关系/转折；仔细阅读每题都能圈出原文证据。", "D01"),
    3: ("基础Ⅱ：讲座讲话 + 推断题", "讲座记录结构词；仔细阅读先定位、再判断，不凭印象。", "D02"),
    4: ("中秋复盘 + 旅行缓冲", "9/25 轻量；9/30–10/1 只维持词感；10/2–10/4 完全留空。", "D03"),
    5: ("旅行后重启：篇章听力 + 长篇匹配", "10/5 完全留空；10/6 起用三天恢复节奏，不补旅行欠账。", "D03"),
    6: ("强化Ⅰ：整套听力控时", "听力 25 题一次做完；阅读优先保证仔细阅读和匹配正确率。", "D04"),
    7: ("强化Ⅱ：高权重讲座 + 阅读控时", "讲座/讲话至少 6/10；阅读 40 分钟内完成并留 2 分钟检查。", "D05"),
    8: ("强化Ⅲ：四项串联 + 首次全真", "工作日完成最后一轮拆题；10/31 做 M01，不暂停、不查词。", "M01"),
    9: ("模考Ⅰ：M01 修复 + M02", "同类错题 48 小时内重做；M02 出版社估分目标 400 左右。", "M02"),
    10: ("模考Ⅱ：运动会减量 + M03", "11/12–14 仅词汇维护；11/15 M03，运动会内容不补课。", "M03"),
    11: ("模考Ⅲ：M03 修复 + M04", "M04 争取 425+；听力和阅读至少一项达到阶段目标。", "M04"),
    12: ("模考Ⅳ：稳定 425 + M05", "减少新技巧；重复正确流程，目标连续第二次接近或超过 425。", "M05"),
    13: ("冲刺Ⅰ：M05 修复 + M06", "M06 争取 450；若未到，最近两套至少稳定 425+。", "M06"),
    14: ("冲刺Ⅱ：封口与减量", "只复习高频错因、模板骨架和听力错句；睡眠优先。", "待确认考试"),
}


PHASES = [
    (date(2026, 9, 7), date(2026, 9, 13), "诊断期"),
    (date(2026, 9, 14), date(2026, 10, 11), "基础期"),
    (date(2026, 10, 12), date(2026, 11, 1), "强化期"),
    (date(2026, 11, 2), date(2026, 11, 29), "模考期"),
    (date(2026, 11, 30), date(2026, 12, 11), "冲刺期"),
    (date(2026, 12, 12), date(2026, 12, 12), "考试占位"),
]


def phase_for(d: date) -> str:
    for start, end, phase in PHASES:
        if start <= d <= end:
            return phase
    raise ValueError(d)


def fmt_task(minutes: int, text: str) -> str:
    if not text:
        return ""
    return f"[{minutes}min] {text}"


def generic_foundation(d: date, week_no: int, paper: str) -> DayPlan:
    wd = d.weekday()
    focus = WEEK_SPECS[week_no][0].split("：", 1)[-1]
    words = fmt_task(25, "不背单词·六级词书：新词 40（周末 50）+ 到期复习；20–25 分钟到点即停")
    if wd == 0:
        return DayPlan(d, week_no, phase_for(d), "工作日", 105, words,
                       fmt_task(40, f"{paper} 长对话 1 组：限时作答→精听 2 轮→听写 6 个转折/问题句"),
                       fmt_task(30, f"{paper} 仔细阅读 1 篇：10 分钟做题 + 20 分钟逐题定位证据"),
                       "", fmt_task(10, "错题写 1 条根因 + 1 条下次动作"),
                       f"{paper}；不背单词；每日英语听力/随书音频", "能指出每道错题的原文证据；精听句可脱稿复述大意")
    if wd == 1:
        return DayPlan(d, week_no, phase_for(d), "工作日", 110, words,
                       fmt_task(30, f"{paper} 篇章听力 1 组：先盲听作答，再逐句回放难句"),
                       fmt_task(35, f"{paper} 长篇匹配：15 分钟限时 + 20 分钟整理同义替换"),
                       fmt_task(20, "真题翻译 3–4 句：先独立写，再对照范文只改语法和搭配"),
                       "", f"{paper}；真题范文", "记录至少 5 组同义替换；翻译改出 3 处可复用表达")
    if wd == 2:
        return DayPlan(d, week_no, phase_for(d), "工作日", 110, words,
                       fmt_task(45, f"{paper} 讲座/讲话 1 组：画结构（主题→分点→例子）并跟读 8 句"),
                       fmt_task(30, f"{paper} 仔细阅读 1 篇：训练{focus}，每题标定位词"),
                       fmt_task(10, "写作：只列 3 段提纲 + 每段 1 个主题句"),
                       "", f"{paper}；随书听力文本", "能用中文复述讲座三层结构；提纲 10 分钟内完成")
    if wd == 3:
        return DayPlan(d, week_no, phase_for(d), "工作日", 115, words,
                       fmt_task(35, f"{paper} 讲座/讲话第 2 组：限时 + 错句精听 + 隔句跟读"),
                       fmt_task(40, f"{paper} 仔细阅读第 2 篇：12 分钟做题 + 28 分钟错项排除"),
                       fmt_task(15, "翻译：复写本周 5 个高频句型，换主题词各造 1 句"),
                       "", f"{paper}；错题记录", "讲座错题二次作答全对；阅读能解释 4 个干扰项为何错")
    if wd == 4:
        return DayPlan(d, week_no, phase_for(d), "工作日", 100, fmt_task(20, "本周词汇到期复习，不加量"),
                       fmt_task(25, "本周精听句盲听复测：不看文本写关键词，错句回放 1 次"),
                       fmt_task(30, f"{paper} 选词填空 1 篇：12 分钟限时；只总结词性和固定搭配"),
                       fmt_task(20, "写作：把周三提纲扩写为开头 + 结尾，各 3–4 句"),
                       fmt_task(5, "标记周末优先复习的 3 个错误"),
                       f"{paper}；本周错题", "精听复测正确率≥80%；形成 1 个可复用开头和结尾")
    if wd == 5:
        return DayPlan(d, week_no, phase_for(d), "周末", 165, fmt_task(25, "新词 50 + 到期复习"),
                       fmt_task(55, f"{paper} 听力 25 题整套限时 30 分钟 + 25 分钟归因"),
                       fmt_task(50, f"{paper} 阅读：长篇匹配 1 篇 + 仔细阅读 1 篇，均计时"),
                       fmt_task(20, "写作：完成 1 个 30 分钟题目的审题、提纲和首段"),
                       fmt_task(15, "将本日错题录入错题记录，安排 48 小时复测"),
                       f"{paper}；计时器；错题记录", "听力全程不停；阅读总用时≤40 分钟；错题全部有根因标签")
    return DayPlan(d, week_no, phase_for(d), "周末", 170, fmt_task(25, "到期词复习 + 本周错词 30 个"),
                   fmt_task(40, "本周听力错句二测：盲听→口头复述→对照文本"),
                   fmt_task(45, "本周阅读错题二测 + 同题型新题 1 篇"),
                   fmt_task(40, "写作/翻译交替：奇数周完整作文，偶数周完整翻译；用范文或 AI 只做批改"),
                   fmt_task(20, "填写每周复盘：完成率、听/读正确率、下周只改 1 件事"),
                   f"{paper}；每周复盘；AI/范文", "二测不看答案；完成周复盘并写出下周唯一优先项")


def generic_strengthen(d: date, week_no: int, paper: str) -> DayPlan:
    wd = d.weekday()
    words = fmt_task(20, "新词 30–35 + 到期复习；把真题错词加入自定义词库")
    if wd == 0:
        return DayPlan(d, week_no, phase_for(d), "工作日", 110, words,
                       fmt_task(40, f"{paper} 听力 25 题：30 分钟限时 + 10 分钟只标错因"),
                       fmt_task(35, f"{paper} 仔细阅读 1 篇 + 长篇匹配半篇，严格计时"),
                       "", fmt_task(15, "挑 2 道最典型错题写证据链"),
                       f"{paper}；计时器", "听力不停顿；阅读每题都有定位词；完成 2 条证据链")
    if wd == 1:
        return DayPlan(d, week_no, phase_for(d), "工作日", 115, words,
                       fmt_task(35, "昨日听力错题精听：只处理错题所在段，听写 8 个关键句"),
                       fmt_task(35, f"{paper} 长篇匹配 1 篇：12–15 分钟 + 同义替换表"),
                       fmt_task(25, "翻译 1 段：25 分钟完成；只检查时态、单复数、主谓一致"),
                       "", f"{paper}；错题记录", "8 个关键句至少 6 句听写正确；翻译无低级语法错")
    if wd == 2:
        return DayPlan(d, week_no, phase_for(d), "工作日", 120, words,
                       fmt_task(40, f"{paper} 讲座/讲话 1 组：记录路标词与层次，复述 60 秒"),
                       fmt_task(45, f"{paper} 仔细阅读 2 篇：22 分钟作答 + 23 分钟订正"),
                       fmt_task(15, "写作：15 分钟完成审题、立场、两条理由和例子"),
                       "", f"{paper}；录音机", "讲座题≥6/10（不足则记录错因）；两篇阅读总作答≤22 分钟")
    if wd == 3:
        return DayPlan(d, week_no, phase_for(d), "工作日", 120, words,
                       fmt_task(25, "隔日盲听复测：周一错题 + 周三讲座，仍错的进入红色清单"),
                       fmt_task(75, f"{paper} 阅读 30 题：40 分钟整套 + 35 分钟精改仔细阅读"),
                       "", "", f"{paper}；红色错题清单", "阅读全套≤40 分钟；仔细阅读正确≥6/10")
    if wd == 4:
        return DayPlan(d, week_no, phase_for(d), "工作日", 110, fmt_task(20, "本周错词复习，不开新词"),
                       fmt_task(25, "红色听力错句跟读 + 变速 0.9×/1.0× 各 1 遍"),
                       fmt_task(30, "本周阅读错题不看答案重做，写出错误选项陷阱"),
                       fmt_task(25, "作文完整写 1 篇（30 分钟题，今天限 25 分钟）"),
                       fmt_task(10, "周末训练清单与设备检查"),
                       "本周错题；真题作文", "错题二测≥80%；作文有明确立场、两条理由、结尾回扣")
    if wd == 5:
        stage_line = {6: "听力≥14/25、阅读≥18/30", 7: "听力≥15/25、阅读≥19/30"}.get(week_no, "听读正确率较上周提升")
        return DayPlan(d, week_no, phase_for(d), "周末", 170, fmt_task(20, "新词 35 + 到期复习"),
                       fmt_task(70, f"{paper} 听力整套 30 分钟 + 精改 40 分钟"),
                       fmt_task(55, f"{paper} 阅读整套 40 分钟 + 重点题复盘 15 分钟"),
                       fmt_task(15, "翻译错句 5 句改写"),
                       fmt_task(10, "记录听力/阅读原始正确数"),
                       f"{paper}；模考记录", f"{stage_line}为阶段线；未达只修最大短板")
    return DayPlan(d, week_no, phase_for(d), "周末", 170, fmt_task(20, "到期词 + 本周错词"),
                   fmt_task(45, "听力错题二测：原音不降速，口头说明为何选该项"),
                   fmt_task(45, "阅读错题二测：重新定位证据并整理 8 组同义替换"),
                   fmt_task(40, "完整翻译或作文 1 篇；按真题范文/评分维度修改一次"),
                   fmt_task(20, "填写每周复盘；选择下周 3 道必复测题"),
                   f"{paper}；每周复盘", "原题二测≥85%；完成一篇修改稿，不抄整篇范文")


def generic_mock_week(d: date, week_no: int, paper: str) -> DayPlan:
    wd = d.weekday()
    previous = {9: "M01", 10: "M02", 11: "M03", 12: "M04"}.get(week_no, "上套模考")
    words = fmt_task(20, "新词 10–20 + 到期复习；11/23 后停止追新词")
    if wd == 0:
        return DayPlan(d, week_no, phase_for(d), "工作日", 115, words,
                       fmt_task(45, f"{previous} 听力错题精听：只练错段，听写 8 句并原速复测"),
                       fmt_task(35, f"{previous} 阅读错题重做：每题写定位句和干扰项陷阱"),
                       "", fmt_task(15, "将最高频错因更新到本周修复清单"),
                       f"{previous}；错题记录", "原题二测≥80%；修复清单最多保留 3 项")
    if wd == 1:
        return DayPlan(d, week_no, phase_for(d), "工作日", 115, words,
                       fmt_task(35, "针对最大听力短板练 1 组同题型：限时→精听→复述"),
                       fmt_task(35, "针对最大阅读短板练 1 组同题型：计时→证据链"),
                       fmt_task(25, f"改写 {previous} 作文/翻译各 1 段，只改表达不重抄全文"),
                       "", "真题错题；范文/AI批改", "两个专项都记录原始正确率；改写稿比原稿少 3 个错误")
    if wd == 2:
        return DayPlan(d, week_no, phase_for(d), "工作日", 120, words,
                       fmt_task(55, "新听力整套 30 分钟 + 25 分钟快速订正，训练连续专注"),
                       fmt_task(35, "仔细阅读 2 篇，22 分钟内完成并标证据"),
                       fmt_task(10, "写作：口头讲清题目立场、理由和例子"),
                       "", "拆题余卷/备用题；计时器", "听力全程不暂停；仔细阅读≥7/10为目标")
    if wd == 3:
        return DayPlan(d, week_no, phase_for(d), "工作日", 115, words,
                       fmt_task(30, "讲座/讲话 1 组：只记结构词和数字/专名，原速复述"),
                       fmt_task(60, "阅读 30 题：40 分钟整套 + 20 分钟只改仔细阅读"),
                       fmt_task(5, "翻译：复习 5 个本周错句"),
                       "", "备用题；错题记录", "阅读≤40 分钟；匹配≥7/10、仔细阅读≥7/10为目标")
    if wd == 4:
        return DayPlan(d, week_no, phase_for(d), "工作日", 95, fmt_task(15, "到期词和错词，不开新词"),
                       fmt_task(20, "本周听力错句原速盲听一遍，禁止过度精听"),
                       fmt_task(25, "阅读同义替换表快速回顾 + 2 道红色错题"),
                       fmt_task(25, "模考作文：列提纲 + 复习 6 个稳妥连接句"),
                       fmt_task(10, "准备铅笔、橡皮、耳机/收音设备和计时环境"),
                       "本周错题；模考用品", "睡前结束，不加练；确认模考卷未看答案")
    if wd == 5:
        return full_mock_day(d, week_no, paper)
    return review_mock_day(d, week_no, paper)


def full_mock_day(d: date, week_no: int, paper: str) -> DayPlan:
    target = {"M01": "≥380", "M02": "≈400", "M03": "≥415", "M04": "≥425", "M05": "≥435", "M06": "≥450（或近两套均≥425）"}.get(paper, "记录基线")
    return DayPlan(d, week_no, phase_for(d), "全真模考", 170,
                   fmt_task(15, "只复习到期词，不开新词"),
                   "", "", "",
                   fmt_task(155, f"{paper} 全真 130 分钟：写作→听力→阅读→翻译，不暂停不查词；25 分钟核对并录入成绩"),
                   f"{paper}（保留卷）；计时器；模考记录",
                   f"严格完成 130 分钟；出版社估分目标 {target}；分数只看趋势，不做简单线性换算")


def review_mock_day(d: date, week_no: int, paper: str) -> DayPlan:
    return DayPlan(d, week_no, phase_for(d), "模考复盘", 170,
                   fmt_task(20, "模考错词 + 到期词，不开新词"),
                   fmt_task(45, f"{paper} 听力错段：先盲听，再看文本听写 8 句，最后原速复测"),
                   fmt_task(45, f"{paper} 阅读错题：重做、标定位句、写干扰项陷阱"),
                   fmt_task(40, f"{paper} 作文/翻译：各选最差 1 段改写；只保留 3 个可复用句型"),
                   fmt_task(20, "填写每周复盘和错题记录；下周只修最大 2 个问题"),
                   f"{paper}；错题记录；每周复盘", "四部分都有复盘产物；48 小时复测日期已写入错题表")


def special_plan(d: date, week_no: int) -> DayPlan | None:
    # Fully protected travel days.
    if date(2026, 10, 2) <= d <= date(2026, 10, 5):
        return DayPlan(d, week_no, phase_for(d), "旅行·完全留空", 0, material="无", success="无需学习，也无需补做；返回后从 10/6 直接继续", status="计划休息")

    if d == date(2026, 9, 7):
        return DayPlan(d, week_no, phase_for(d), "工作日", 105,
                       fmt_task(25, "安装/选定不背单词，选择六级词书；做 10 分钟词汇熟悉度抽测后开始新词 30"),
                       fmt_task(20, "阅读 CET-6 听力题型与答题顺序；试听随书音频，确认设备"),
                       fmt_task(20, "阅读题型与时间分配：选词 8 分钟、匹配 12 分钟、仔细阅读 20 分钟"),
                       fmt_task(20, "看 1 道作文和 1 道翻译要求，只做审题不看范文"),
                       fmt_task(20, "建立真题编号 F0/D01–D05/M01–M06；填写总览基础分"),
                       "本工作簿；真题册说明页；不背单词", "所有工具可用；12 套真题已编号；未提前看保留模考卷答案")
    if d == date(2026, 9, 8):
        return DayPlan(d, week_no, phase_for(d), "工作日", 110,
                       fmt_task(25, "新词 35 + 到期复习"),
                       fmt_task(45, "用真题册例题做 1 组听力：限时→标错因→精听 6 个难句"),
                       fmt_task(20, "把听力错因分为听不出、词不懂、定位错、走神"),
                       "", fmt_task(20, "记录耳机/音量、最易走神的时段和应对动作"),
                       "真题册例题（非 F0）；随书音频", "至少找到 2 个主要听力问题；6 个难句可原速跟读")
    if d == date(2026, 9, 9):
        return DayPlan(d, week_no, phase_for(d), "工作日", 110,
                       fmt_task(25, "新词 35 + 到期复习"),
                       "", fmt_task(60, "用真题册例题做：匹配 1 篇 + 仔细阅读 1 篇；逐题圈定位句和同义替换"),
                       "", fmt_task(25, "把阅读错因分为词汇、定位、长难句、逻辑、超时"),
                       "真题册例题（非 F0）", "完成阅读时间记录；每道错题都能归入一个错因")
    if d == date(2026, 9, 10):
        return DayPlan(d, week_no, phase_for(d), "工作日", 110,
                       fmt_task(25, "新词 35 + 到期复习"),
                       "", "", fmt_task(60, "作文 30 分钟 + 翻译 30 分钟，均独立完成后再看范文"),
                       fmt_task(25, "各挑 3 个问题：审题/结构/语法/搭配；不抄整篇范文"),
                       "真题册例题；计时器", "两项都在时限内完成；得到各 3 个可执行改进点")
    if d == date(2026, 9, 11):
        return DayPlan(d, week_no, phase_for(d), "工作日", 95,
                       fmt_task(20, "只做本周到期复习"),
                       fmt_task(20, "重听 9/8 的 6 个难句，不看文本复述"),
                       fmt_task(20, "重做 9/9 的错题，不看答案"),
                       fmt_task(15, "整理明天 F0 模考的作文/翻译时间节点"),
                       fmt_task(20, "写下基线预测与三条模考纪律；今晚减量"),
                       "本周错题；F0（不得翻阅）", "设备就绪；错题二测完成；今晚不加练")
    if d == date(2026, 9, 12):
        return full_mock_day(d, week_no, "F0")
    if d == date(2026, 9, 13):
        return review_mock_day(d, week_no, "F0")

    if d == date(2026, 9, 20):
        return DayPlan(d, week_no, phase_for(d), "调休上课日", 110,
                       fmt_task(20, "到期词复习；按工作日强度，不加新词"),
                       fmt_task(35, "D01 本周错句盲听 + 跟读；只修最常见错因"),
                       fmt_task(30, "D01 仔细阅读错题二测，重新标证据"),
                       fmt_task(15, "完整翻译只改写最差 3 句"),
                       fmt_task(10, "简版周复盘：下周唯一重点"),
                       "D01；国务院调休安排", "按调休上课日完成 110 分钟；不补成周末大课")

    if d == date(2026, 9, 25):
        return DayPlan(d, week_no, phase_for(d), "中秋·轻量", 35,
                       fmt_task(20, "到期词 + 本周错词；不加新词"),
                       fmt_task(15, "D02 精听句盲听一遍，能复述大意即可"),
                       material="不背单词；D02 音频", success="35 分钟到点结束；不安排补课")
    if d == date(2026, 9, 26):
        return DayPlan(d, week_no, phase_for(d), "中秋假期·弹性", 125,
                       fmt_task(20, "新词 30 + 到期复习"),
                       fmt_task(45, "D02 听力整套限时 + 只改最有代表性的 5 题"),
                       fmt_task(35, "D02 仔细阅读 2 篇，计时并标证据"),
                       fmt_task(15, "翻译错句改写 3 句"),
                       fmt_task(10, "未完成可直接跳过，不顺延"),
                       "D02；计时器", "完成核心听读即可；总时长不超过 125 分钟")
    if d == date(2026, 9, 27):
        return DayPlan(d, week_no, phase_for(d), "中秋假期·弹性", 145,
                       fmt_task(20, "到期词 + 本周错词"),
                       fmt_task(35, "D02 听力错题二测"),
                       fmt_task(40, "D02 阅读错题二测 + 5 组同义替换"),
                       fmt_task(30, "完整作文 1 篇，按 30 分钟完成"),
                       fmt_task(20, "周复盘 + 旅行前任务封箱"),
                       "D02；每周复盘", "本周结清，不把任务带入旅行")

    if d == date(2026, 9, 30):
        return DayPlan(d, week_no, phase_for(d), "旅行·仅词汇", 20,
                       fmt_task(20, "只做 App 到期词；不加新词，不补任何听读"),
                       material="不背单词", success="20 分钟到点停止", status="未开始")
    if d == date(2026, 10, 1):
        return DayPlan(d, week_no, phase_for(d), "旅行·仅词汇", 15,
                       fmt_task(15, "只复习 App 到期词；若行程紧可直接跳过且不补"),
                       material="不背单词", success="最多 15 分钟；不形成欠账", status="未开始")

    if d == date(2026, 10, 6):
        return DayPlan(d, week_no, phase_for(d), "旅行后恢复", 80,
                       fmt_task(20, "到期词复习；只加 20 个新词"),
                       fmt_task(30, "D03 选 1 组长对话：限时 + 只精听 4 个难句"),
                       fmt_task(25, "D03 仔细阅读 1 篇：12 分钟做题 + 13 分钟定位"),
                       review=fmt_task(5, "写下旅行后第一阻力，不补旧任务"),
                       material="D03；不背单词", success="完成 80 分钟即恢复成功；不追赶")
    if d == date(2026, 10, 7):
        return DayPlan(d, week_no, phase_for(d), "旅行后恢复", 100,
                       fmt_task(20, "新词 25 + 到期复习"),
                       fmt_task(35, "D03 篇章听力 1 组：限时 + 6 句跟读"),
                       fmt_task(30, "D03 长篇匹配 1 篇：15 分钟作答 + 15 分钟同义替换"),
                       fmt_task(10, "作文列 3 段提纲"), fmt_task(5, "确认 10/8 恢复正常"),
                       "D03；随书音频", "听读各完成一项；不超过 100 分钟")
    if d == date(2026, 10, 10):
        return DayPlan(d, week_no, phase_for(d), "调休上课日", 105,
                       fmt_task(20, "到期词 + 错词，不加量"),
                       fmt_task(35, "D03 听力整套后只精改错题所在段"),
                       fmt_task(30, "D03 仔细阅读 1 篇 + 证据定位"),
                       fmt_task(10, "翻译错句 3 句"), fmt_task(10, "记录正确率"),
                       "D03；国务院调休安排", "按工作日完成；不因周六排 3 小时")

    # First full mock and later full mocks/reviews.
    mock_dates = {
        date(2026, 10, 31): "M01",
        date(2026, 11, 7): "M02",
        date(2026, 11, 15): "M03",
        date(2026, 11, 21): "M04",
        date(2026, 11, 28): "M05",
        date(2026, 12, 5): "M06",
    }
    if d in mock_dates:
        return full_mock_day(d, week_no, mock_dates[d])
    review_dates = {
        date(2026, 11, 1): "M01",
        date(2026, 11, 8): "M02",
        date(2026, 11, 22): "M04",
        date(2026, 11, 29): "M05",
        date(2026, 12, 6): "M06",
    }
    if d in review_dates:
        return review_mock_day(d, week_no, review_dates[d])

    # Sports meeting: vocabulary only, no catch-up.
    if date(2026, 11, 12) <= d <= date(2026, 11, 14):
        minutes = 20 if d != date(2026, 11, 14) else 15
        return DayPlan(d, week_no, phase_for(d), "运动会·轻量", minutes,
                       fmt_task(minutes, "只做 App 到期词和最近模考错词；不加新词"),
                       material="不背单词；M02 错词", success=f"最多 {minutes} 分钟；运动会任务不补课")

    # Final sprint, all individually tuned.
    final_days: Dict[date, DayPlan] = {}
    def final(day_: date, mins: int, wt: Tuple[int, str], li: Tuple[int, str], rd: Tuple[int, str], wr: Tuple[int, str], rv: Tuple[int, str], success: str) -> DayPlan:
        return DayPlan(day_, 14 if day_ >= date(2026, 12, 7) else 13, phase_for(day_), "考前减量" if day_ >= date(2026, 12, 7) else "工作日", mins,
                       fmt_task(*wt) if wt[0] else "", fmt_task(*li) if li[0] else "", fmt_task(*rd) if rd[0] else "", fmt_task(*wr) if wr[0] else "", fmt_task(*rv) if rv[0] else "",
                       "M05/M06 错题；红色清单", success)
    final_days[date(2026, 11, 30)] = final(date(2026, 11, 30), 115, (15, "错词和到期词，不开新词"), (40, "M05 听力错段精听 + 原速二测"), (35, "M05 阅读错题重做并标证据"), (15, "M05 作文最差一段改写"), (10, "只保留两个最大问题"), "M05 原题二测≥80%；修复项≤2")
    final_days[date(2026, 12, 1)] = final(date(2026, 12, 1), 110, (15, "错词复习"), (35, "最大听力短板 1 组，限时后只精听错段"), (35, "最大阅读短板 1 组，严格计时"), (20, "翻译 1 段，检查三类低级错"), (5, "记录正确率"), "两项专项均完成并记录正确率")
    final_days[date(2026, 12, 2)] = final(date(2026, 12, 2), 110, (15, "错词复习"), (45, "听力整套 30 分钟 + 快速订正 15 分钟"), (35, "仔细阅读 2 篇，22 分钟内作答"), (10, "作文口头提纲"), (5, "标记 M06 前注意点"), "听力不停顿；阅读≥7/10")
    final_days[date(2026, 12, 3)] = final(date(2026, 12, 3), 105, (15, "错词复习"), (25, "讲座/讲话 1 组原速复述"), (45, "阅读整套 40 分钟 + 5 分钟记录"), (15, "翻译复习 5 个稳妥句型"), (5, "检查时间分配"), "阅读≤40 分钟；不引入新技巧")
    final_days[date(2026, 12, 4)] = final(date(2026, 12, 4), 75, (15, "到期词，不开新词"), (15, "红色错句盲听一遍"), (15, "红色阅读错题 3 道"), (20, "M06 作文提纲 + 连接句"), (10, "准备模考环境，提前结束"), "75 分钟到点停止；保证睡眠")
    final_days[date(2026, 12, 7)] = final(date(2026, 12, 7), 100, (15, "M06 错词"), (35, "M06 最难 3 段原速二测"), (30, "M06 仔细阅读错题重做"), (10, "作文最差一段改写"), (10, "确认最终两项提醒"), "只修已知错误，不开新材料")
    final_days[date(2026, 12, 8)] = final(date(2026, 12, 8), 100, (15, "错词复习"), (30, "听力半套 15 分钟 + 快速核对"), (30, "仔细阅读 1 篇 + 匹配半篇，合计 25 分钟内"), (20, "翻译 1 段，限时 20 分钟"), (5, "记录节奏"), "保持手感；所有任务一次完成不重刷")
    final_days[date(2026, 12, 9)] = final(date(2026, 12, 9), 90, (15, "错词复习"), (25, "讲座错句与数字/专名清单"), (25, "阅读同义替换清单 + 2 道错题"), (20, "作文 3 类题目各列 1 个提纲"), (5, "背准而非背多"), "三个提纲均可 5 分钟内完成")
    final_days[date(2026, 12, 10)] = final(date(2026, 12, 10), 75, (15, "只复习错词"), (20, "听力红色清单原速一遍"), (20, "阅读红色清单一遍"), (15, "翻译常错语法 + 5 个连接句"), (5, "准考证/证件/设备清单"), "不做整套题；用品清单已确认")
    final_days[date(2026, 12, 11)] = final(date(2026, 12, 11), 40, (10, "只看熟悉错词"), (10, "听 1 段已做过的原速材料"), (0, ""), (10, "看自己的作文骨架与翻译错句"), (10, "确认时间地点，23:00 前睡"), "40 分钟后彻底停止；不估分、不刷陌生题")
    if d in final_days:
        return final_days[d]

    if d == PLACEHOLDER_EXAM_DATE:
        return DayPlan(d, week_no, phase_for(d), "暂定考试日", 0,
                       review="暂按 CET-6 15:00–17:25 预留；正式日期和时段以教育部教育考试院公告/准考证为准。若当天不是考试：把本日改为 M06 复测或轻量全真，并将冲刺周顺延。",
                       material="准考证、身份证、2B 铅笔、黑色签字笔、耳机/收音设备（以学校通知为准）",
                       success="考试日不安排额外学习；该日期为排程占位，待官方确认", status="待确认")
    return None


def build_daily_plans() -> List[DayPlan]:
    plans: List[DayPlan] = []
    d = START_DATE
    while d <= PLACEHOLDER_EXAM_DATE:
        week_no = ((d - START_DATE).days // 7) + 1
        special = special_plan(d, week_no)
        if special is not None:
            plans.append(special)
        else:
            _, _, paper = WEEK_SPECS[week_no]
            if week_no <= 5:
                plans.append(generic_foundation(d, week_no, paper))
            elif week_no <= 8:
                plans.append(generic_strengthen(d, week_no, paper))
            elif week_no <= 12:
                plans.append(generic_mock_week(d, week_no, paper))
            else:
                # Week 13 generic mock logic only applies before custom final days; Sunday is handled above.
                plans.append(generic_mock_week(d, week_no, paper))
        d += timedelta(days=1)
    return plans


def thin_border(color: str = LINE) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def style_title(ws, title: str, subtitle: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    c = ws.cell(1, 1, title)
    c.font = Font(name="Microsoft YaHei", size=20, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    c = ws.cell(2, 1, subtitle)
    c.font = Font(name="Microsoft YaHei", size=10, color=MUTED)
    c.fill = PatternFill("solid", fgColor=OFFWHITE)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 25


def section_header(ws, row: int, text: str, start_col: int, end_col: int, fill: str = TEAL) -> None:
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row, start_col, text)
    c.font = Font(name="Microsoft YaHei", size=12, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 25


def table_header(ws, row: int, start_col: int, values: List[str], fill: str = NAVY) -> None:
    for idx, value in enumerate(values, start_col):
        c = ws.cell(row, idx, value)
        c.font = Font(name="Microsoft YaHei", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 30


def set_all_font(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.font.name != "Microsoft YaHei":
                cell.font = Font(name="Microsoft YaHei", size=cell.font.sz or 10, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)


def build_overview(wb: Workbook, plans: List[DayPlan]) -> None:
    ws = wb.active
    ws.title = "总览"
    ws.sheet_view.showGridLines = False
    style_title(ws, "2026 年 12 月 CET-6 425+ 备考计划", "初版 v1.0｜按 450 能力构建容错｜莆田学院｜生成于 2026-08-27", 12)

    section_header(ws, 4, "一眼看懂：当前差距与目标", 1, 6)
    table_header(ws, 5, 1, ["模块", "当前分", "训练目标", "需提升", "策略", "优先级"])
    score_rows = [
        ("听力", 78, 150, "=C6-B6", "真题原音；讲座/讲话和错句精听优先", "核心"),
        ("阅读", 86, 165, "=C7-B7", "仔细阅读 + 长篇匹配优先；选词填空低投入", "核心"),
        ("写作+翻译", 117, 135, "=C8-B8", "每周 1 篇完整输出 + 改写；维持优势", "维持提分"),
        ("总分", 281, 450, "=C9-B9", "以两次连续 425+、至少一次 450 左右为考前信号", "目标"),
    ]
    for r_idx, values in enumerate(score_rows, 6):
        for c_idx, value in enumerate(values, 1):
            c = ws.cell(r_idx, c_idx, value)
            c.font = Font(name="Microsoft YaHei", size=10, bold=(r_idx == 9), color=INK)
            c.fill = PatternFill("solid", fgColor=PALE_BLUE if r_idx < 9 else PALE_GREEN)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r_idx].height = 34

    section_header(ws, 11, "考试与校历状态", 1, 6, AMBER)
    status_rows = [
        ("正式开练", START_DATE, "已确认", "校历 9/7 开学"),
        ("中秋节", date(2026, 9, 25), "已确认", "9/25 轻量；国务院假期 9/25–27"),
        ("新加坡旅行", "2026-09-30 至 2026-10-05", "用户确认", "9/30–10/1 仅词汇；10/2–10/5 完全留空"),
        ("运动会", "2026-11-12 至 2026-11-14", "校历可明确识别", "三天只留词汇维护；不补课"),
        ("CET-6 笔试", PLACEHOLDER_EXAM_DATE, "待官方确认", "仅作倒排占位；暂按 15:00–17:25 预留"),
    ]
    table_header(ws, 12, 1, ["事项", "日期", "状态", "计划处理"], fill=AMBER)
    for r_idx, values in enumerate(status_rows, 13):
        for c_idx, value in enumerate(values, 1):
            c = ws.cell(r_idx, c_idx, value)
            c.font = Font(name="Microsoft YaHei", size=10, color=INK, bold=(values[2] == "待官方确认"))
            c.fill = PatternFill("solid", fgColor=PALE_AMBER if values[2] == "待官方确认" else WHITE)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(value, date):
                c.number_format = "yyyy-mm-dd"
        ws.row_dimensions[r_idx].height = 32

    section_header(ws, 19, "阶段路线", 1, 6)
    table_header(ws, 20, 1, ["阶段", "日期", "主目标", "听力", "阅读", "写译"])
    phase_rows = [
        ("诊断期", "9/7–9/13", "完成 F0 基线并分类错因", "找主要失分点", "找主要失分点", "完成两项限时样本"),
        ("基础期", "9/14–10/11", "把做题流程固定下来", "长对话/篇章/讲座逐类", "证据定位与同义替换", "每周 1 次完整输出"),
        ("强化期", "10/12–11/1", "整模块计时 + 首次全真", "14–15/25", "18–19/30", "30 分钟内完成"),
        ("模考期", "11/2–11/29", "M02–M05 每周一套", "向 16–17/25 靠拢", "向 20–21/30 靠拢", "改写而非抄范文"),
        ("冲刺期", "11/30–12/11", "M06 + 错题封口 + 减量", "保持原速手感", "稳定时间分配", "只复习个人骨架"),
    ]
    for r_idx, values in enumerate(phase_rows, 21):
        for c_idx, value in enumerate(values, 1):
            c = ws.cell(r_idx, c_idx, value)
            c.font = Font(name="Microsoft YaHei", size=9, color=INK)
            c.fill = PatternFill("solid", fgColor=WHITE if r_idx % 2 else PALE_BLUE)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r_idx].height = 36

    section_header(ws, 27, "执行数据（随每日计划更新）", 1, 6, TEAL_DARK)
    dashboard = [
        ("计划学习天数", "=COUNTIF('每日计划'!$F:$F,\">0\")", "天"),
        ("已完成天数", "=COUNTIF('每日计划'!$N:$N,\"已完成\")", "天"),
        ("总体完成率", "=IFERROR(B29/B28,0)", "%"),
        ("计划总时长", "=SUM('每日计划'!$F:$F)/60", "小时"),
        ("实际总时长", "=SUM('每日计划'!$O:$O)/60", "小时"),
        ("最近模考估分", "=IFERROR(LOOKUP(2,1/('模考记录'!$H:$H<>\"\"),'模考记录'!$H:$H),\"待填写\")", "分"),
    ]
    table_header(ws, 28, 1, ["指标", "当前值", "单位"], fill=TEAL_DARK)
    # Header occupies row 28, so data starts row 29. Formulas are adjusted below.
    dashboard[2] = ("总体完成率", "=IFERROR(B30/B29,0)", "%")
    for r_idx, values in enumerate(dashboard, 29):
        for c_idx, value in enumerate(values, 1):
            c = ws.cell(r_idx, c_idx, value)
            c.font = Font(name="Microsoft YaHei", size=10, bold=(c_idx == 2), color=INK)
            c.fill = PatternFill("solid", fgColor=PALE_GREEN)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center")
        if values[2] == "%":
            ws.cell(r_idx, 2).number_format = "0%"
        elif values[2] == "小时":
            ws.cell(r_idx, 2).number_format = "0.0"

    # Score comparison chart.
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "当前分 vs 训练目标"
    chart.y_axis.title = "分数"
    chart.x_axis.title = "模块"
    chart.height = 7.2
    chart.width = 12.5
    data = Reference(ws, min_col=2, max_col=3, min_row=5, max_row=8)
    cats = Reference(ws, min_col=1, min_row=6, max_row=8)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend.position = "b"
    ws.add_chart(chart, "H4")

    # Daily operating rules.
    section_header(ws, 27, "三条执行规则", 8, 12, PURPLE)
    rules = [
        "1. 未完成不整体顺延：只把当天最重要的一项放到周日，最多带 1 项欠账。",
        "2. 最低可执行版（忙到爆时）：10 分钟到期词 + 15 分钟错句盲听 + 15 分钟仔细阅读证据定位，共 40 分钟。",
        "3. 真题顺序：F0/D01–D05 可拆；M01–M06 必须保留到指定模考日，提前不看答案。",
        "4. 估分只能看趋势：CET 采用常模转换，原始正确数不能简单线性换算成 710 分制。",
        "5. 若正式考试晚于 12/12：将 12/12 改为轻量全真/复测，并重复最后冲刺周；若提前，则从考前 6 天直接执行第 14 周。",
    ]
    for idx, text in enumerate(rules, 28):
        ws.merge_cells(start_row=idx, start_column=8, end_row=idx, end_column=12)
        c = ws.cell(idx, 8, text)
        c.font = Font(name="Microsoft YaHei", size=10, color=INK)
        c.fill = PatternFill("solid", fgColor=PALE_PURPLE if idx % 2 == 0 else WHITE)
        c.border = thin_border()
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[idx].height = 37

    section_header(ws, 36, "官方状态说明（务必保留）", 1, 12, AMBER)
    ws.merge_cells("A37:L38")
    c = ws["A37"]
    c.value = (
        "截至 2026-08-27，教育部教育考试院 CET 考试动态与报名网仍未发布 2026 年下半年正式报名/考试公告。"
        "本计划以 2026-12-12 为倒排占位日；上海市年度考试计划曾列该日，但它不是本次中央正式公告，且年度计划可能调整。"
        "CET-6 暂按近次官方时段 15:00–17:25 预留，最终以教育部教育考试院公告、莆田学院通知和准考证为准。"
    )
    c.font = Font(name="Microsoft YaHei", size=10, bold=True, color=RED)
    c.fill = PatternFill("solid", fgColor=PALE_AMBER)
    c.border = thin_border(AMBER)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[37].height = 32
    ws.row_dimensions[38].height = 32

    for col, width in {"A":16, "B":16, "C":16, "D":16, "E":31, "F":14, "G":3, "H":17, "I":17, "J":17, "K":17, "L":17}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"
    ws.print_area = "A1:L38"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 2
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_daily_sheet(wb: Workbook, plans: List[DayPlan]) -> None:
    ws = wb.create_sheet("每日计划")
    ws.sheet_view.showGridLines = False
    headers = ["日期", "周次", "星期", "阶段", "日型", "计划(min)", "单词", "听力", "阅读", "写作/翻译", "复盘/输出", "资料/编号", "当日完成标准", "状态", "实际(min)", "自评(1-5)", "备注"]
    for c_idx, value in enumerate(headers, 1):
        c = ws.cell(1, c_idx, value)
        c.font = Font(name="Microsoft YaHei", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[1].height = 34

    fill_map = {
        "旅行·完全留空": GREY,
        "旅行·仅词汇": PALE_AMBER,
        "中秋·轻量": PALE_AMBER,
        "中秋假期·弹性": "FFF9E9",
        "运动会·轻量": PALE_PURPLE,
        "全真模考": PALE_GREEN,
        "模考复盘": PALE_BLUE,
        "暂定考试日": PALE_RED,
        "考前减量": "F3F4F6",
        "调休上课日": "EAF0F8",
        "旅行后恢复": "F0F7F2",
    }
    for r_idx, plan in enumerate(plans, 2):
        values = [
            plan.day, plan.week_no, WEEKDAY_CN[plan.day.weekday()], plan.phase, plan.day_type,
            plan.planned_minutes, plan.word, plan.listening, plan.reading, plan.writing_translation,
            plan.review, plan.material, plan.success, plan.status, None, None, None,
        ]
        row_fill = fill_map.get(plan.day_type, (OFFWHITE if plan.day.weekday() >= 5 else WHITE))
        for c_idx, value in enumerate(values, 1):
            c = ws.cell(r_idx, c_idx, value)
            c.font = Font(name="Microsoft YaHei", size=9, color=INK, bold=(c_idx in (1, 5, 6)))
            c.fill = PatternFill("solid", fgColor=row_fill)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center" if c_idx in (1, 2, 3, 4, 5, 6, 14, 15, 16) else "left", vertical="top", wrap_text=True)
        ws.cell(r_idx, 1).number_format = "yyyy-mm-dd"
        ws.row_dimensions[r_idx].height = 72 if plan.planned_minutes else 48

    end_row = len(plans) + 1
    tab = XLTable(displayName="DailyPlan", ref=f"A1:Q{end_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    ws.add_table(tab)

    status_dv = DataValidation(type="list", formula1='"未开始,进行中,已完成,计划休息,跳过,补做,待确认"', allow_blank=False)
    rating_dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=True)
    minutes_dv = DataValidation(type="whole", operator="between", formula1="0", formula2="600", allow_blank=True)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(rating_dv)
    ws.add_data_validation(minutes_dv)
    status_dv.add(f"N2:N{end_row}")
    minutes_dv.add(f"O2:O{end_row}")
    rating_dv.add(f"P2:P{end_row}")

    ws.conditional_formatting.add(f"N2:N{end_row}", FormulaRule(formula=["$N2=\"已完成\""], fill=PatternFill("solid", fgColor=PALE_GREEN)))
    ws.conditional_formatting.add(f"N2:N{end_row}", FormulaRule(formula=["$N2=\"进行中\""], fill=PatternFill("solid", fgColor=PALE_AMBER)))
    ws.conditional_formatting.add(f"N2:N{end_row}", FormulaRule(formula=["OR($N2=\"跳过\",$N2=\"补做\")"], fill=PatternFill("solid", fgColor=PALE_RED)))
    ws.conditional_formatting.add(f"O2:O{end_row}", CellIsRule(operator="greaterThan", formula=["$F2*1.2"], fill=PatternFill("solid", fgColor=PALE_AMBER)))

    widths = [12, 7, 8, 11, 17, 10, 35, 44, 44, 40, 40, 28, 42, 12, 11, 11, 32]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "G2"
    # The Excel table already supplies filters. A second overlapping sheet-level
    # AutoFilter makes some desktop Excel builds reject the workbook.
    ws.print_title_rows = "1:1"
    ws.print_area = f"A1:Q{end_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = "第 &P / &N 页"
    ws.oddFooter.right.text = "CET-6 计划 v1.0"


def build_resources_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("资料清单")
    ws.sheet_view.showGridLines = False
    style_title(ws, "资料清单：少买、用透", "核心原则：只买一套真题；电子工具各司其职，不同时开多个同类 App。价格为大致区间，以购买时为准。", 8)
    headers = ["类别", "资料/工具", "优先级", "预计费用", "怎么用", "频率", "替代/备注", "链接"]
    table_header(ws, 4, 1, headers)
    rows = [
        ("纸质·核心", "最新六级真题详解（首选星火；新东方同类最新版也可，二选一）", "必备", "约 35–60 元", "购买时确认：至少 12 套完整真题、原版音频、听力原文、逐题解析；按 F0/D01–D05/M01–M06 编号", "全程", "不要同时买两套；书名随版本略变，以功能为准", ""),
        ("纸质·可选", "新东方六级词汇乱序版/同类高频词汇书", "可选", "约 30–45 元", "只在不喜欢 App 时使用；每天按时间上限背，不追页码", "每日 20–25min", "已用 App 就不买", ""),
        ("电子·词汇", "不背单词（六级词书）", "必备", "基础功能可免费", "前期新词 35–40 + 到期复习；后期停新词，只复习真题错词", "每日", "墨墨/百词斩可替代，但只能选一个", "https://www.bbdc.cn/"),
        ("电子·真题音频", "所购真题册随书音频/小程序", "必备", "随书", "始终以真题原音为主；下载离线，按句回放", "每周 5–6 次", "音频必须与试卷版本对应", ""),
        ("电子·精听", "每日英语听力", "推荐", "基础功能可免费", "导入/搜索真题音频，使用句子回放、变速和文本对照；不做整篇逐字听写", "每周 4–5 次", "随书播放器若能逐句回放也够用", "https://dict.eudic.net/ting"),
        ("电子·写译批改", "当前使用的 ChatGPT/Codex 或学校批改网", "推荐", "已有服务/校内资源", "先独立限时写，再要求按 CET-6 维度指出 3–5 个问题并给改写理由；禁止直接代写", "每周 1–2 次", "真题范文自查也可", ""),
        ("电子·错题", "本工作簿：错题记录 + 模考记录", "必备", "免费", "只记高价值错题：错因、证据、下次动作、48 小时/7 天复测", "每次训练后", "不另装笔记 App，减少切换", ""),
        ("电子·补充听力", "BBC Learning English · 6 Minute English", "可选", "免费", "只作旅行/轻量日补充；真题完成后再听，不替代真题", "每周≤1 次", "听懂主旨即可，不逐字查词", "https://www.bbc.co.uk/learningenglish/english/features/6-minute-english"),
        ("官方信息", "全国大学英语四六级考试官网/报名网", "必备", "免费", "核对报名、考试日期、准考证与成绩；9 月起每周检查一次学校通知即可", "每周 1 次", "不要以培训机构日历代替官方公告", NEEA_NEWS_URL),
        ("硬件", "现有耳机/收音设备 + 计时器", "必备", "0 元（已有）", "精听、模考使用同一设备；提前确认学校听力接收方式", "全程", "考试用品以莆田学院通知为准", ""),
    ]
    for r_idx, row in enumerate(rows, 5):
        for c_idx, value in enumerate(row, 1):
            c = ws.cell(r_idx, c_idx, value)
            c.font = Font(name="Microsoft YaHei", size=9, color=INK, bold=(c_idx == 2 and row[2] == "必备"))
            c.fill = PatternFill("solid", fgColor=PALE_GREEN if row[2] == "必备" else (PALE_BLUE if row[2] == "推荐" else WHITE))
            c.border = thin_border()
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if c_idx == 8 and value:
                c.hyperlink = value
                c.style = "Hyperlink"
        ws.row_dimensions[r_idx].height = 58

    section_header(ws, 17, "推荐购买结论", 1, 8, TEAL_DARK)
    ws.merge_cells("A18:H19")
    ws["A18"] = (
        "性价比方案：纸质只买 1 套最新版六级真题（约 35–60 元）。词汇用不背单词，精听用随书音频 + 每日英语听力，"
        "写译用真题范文或现有 AI 做纠错，本工作簿直接承担错题与模考记录。若确实偏爱纸质背词，再加 1 本词汇书；不建议购买阅读、听力、作文三本专项书。"
    )
    ws["A18"].font = Font(name="Microsoft YaHei", size=11, bold=True, color=INK)
    ws["A18"].fill = PatternFill("solid", fgColor=PALE_GREEN)
    ws["A18"].border = thin_border(GREEN)
    ws["A18"].alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[18].height = 32
    ws.row_dimensions[19].height = 32

    widths = [15, 34, 11, 15, 52, 18, 42, 38]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A5"
    ws.print_area = "A1:H19"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_paper_map_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("真题映射")
    ws.sheet_view.showGridLines = False
    style_title(ws, "真题编号与使用顺序", "先按用途编号，再把所购版本的实际年份/套次填入。M01–M06 在指定日期前不得看答案。", 9)
    headers = ["编号", "用途", "建议选卷", "实际年份/套次（购买后填）", "是否看过答案", "计划日期", "完成日期", "估分/正确率", "备注"]
    table_header(ws, 4, 1, headers)
    planned_dates = {
        "F0": date(2026, 9, 12), "M01": date(2026, 10, 31), "M02": date(2026, 11, 7),
        "M03": date(2026, 11, 15), "M04": date(2026, 11, 21), "M05": date(2026, 11, 28), "M06": date(2026, 12, 5),
    }
    rows = [("F0", "完整基线", "最旧且完整的一套", "", "否", planned_dates["F0"], "", "", "一次完成，保留真实基线")]
    for i in range(1, 6):
        rows.append((f"D{i:02d}", "拆分专项", f"从旧到新第 {i+1} 套", "", "可拆", "按每日计划", "", "", "可分听力/阅读/写译训练"))
    for i in range(1, 7):
        code = f"M{i:02d}"
        rows.append((code, "全真模考", f"最新六套中从旧到新第 {i} 套", "", "否", planned_dates[code], "", "", "指定日期前不翻答案/听力原文"))
    for r_idx, row in enumerate(rows, 5):
        fill = PALE_GREEN if row[1] == "全真模考" else (PALE_BLUE if row[1] == "拆分专项" else PALE_AMBER)
        for c_idx, value in enumerate(row, 1):
            c = ws.cell(r_idx, c_idx, value)
            c.font = Font(name="Microsoft YaHei", size=9, bold=(c_idx == 1), color=INK)
            c.fill = PatternFill("solid", fgColor=fill)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(value, date):
                c.number_format = "yyyy-mm-dd"
        ws.row_dimensions[r_idx].height = 36
    end_row = 4 + len(rows)
    dv = DataValidation(type="list", formula1='"否,是,可拆"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"E5:E{end_row}")
    widths = [11, 16, 28, 30, 16, 17, 17, 18, 38]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A5"
    ws.print_area = f"A1:I{end_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_weekly_review_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("每周复盘")
    ws.sheet_view.showGridLines = False
    style_title(ws, "每周复盘", "每周日用 15–20 分钟填写。完成率低不等于失败：只找一个最值得调整的变量。", 15)
    headers = ["周次", "开始", "结束", "本周主题", "计划(min)", "实际(min)", "计划天数", "完成天数", "完成率", "听力正确/25", "阅读正确/30", "模考估分", "最大问题", "下周唯一调整", "状态"]
    table_header(ws, 4, 1, headers)
    for week_no in range(1, 15):
        r = week_no + 4
        start = START_DATE + timedelta(days=(week_no - 1) * 7)
        end = min(start + timedelta(days=6), PLACEHOLDER_EXAM_DATE)
        theme, _, _ = WEEK_SPECS[week_no]
        values = [
            week_no, start, end, theme,
            f'=SUMIFS(\'每日计划\'!$F:$F,\'每日计划\'!$A:$A,\">=\"&B{r},\'每日计划\'!$A:$A,\"<=\"&C{r})',
            f'=SUMIFS(\'每日计划\'!$O:$O,\'每日计划\'!$A:$A,\">=\"&B{r},\'每日计划\'!$A:$A,\"<=\"&C{r})',
            f'=COUNTIFS(\'每日计划\'!$A:$A,\">=\"&B{r},\'每日计划\'!$A:$A,\"<=\"&C{r},\'每日计划\'!$F:$F,\">0\")',
            f'=COUNTIFS(\'每日计划\'!$A:$A,\">=\"&B{r},\'每日计划\'!$A:$A,\"<=\"&C{r},\'每日计划\'!$N:$N,\"已完成\")',
            f'=IFERROR(H{r}/G{r},0)', None, None, None, None, None, "未复盘",
        ]
        for c_idx, value in enumerate(values, 1):
            c = ws.cell(r, c_idx, value)
            c.font = Font(name="Microsoft YaHei", size=9, color=INK)
            c.fill = PatternFill("solid", fgColor=PALE_BLUE if week_no % 2 == 0 else WHITE)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center" if c_idx <= 12 else "left")
        ws.cell(r, 2).number_format = "m/d"
        ws.cell(r, 3).number_format = "m/d"
        ws.cell(r, 9).number_format = "0%"
        ws.row_dimensions[r].height = 48

    dv_status = DataValidation(type="list", formula1='"未复盘,已复盘,需降量,需补强"', allow_blank=False)
    dv_listen = DataValidation(type="whole", operator="between", formula1="0", formula2="25", allow_blank=True)
    dv_read = DataValidation(type="whole", operator="between", formula1="0", formula2="30", allow_blank=True)
    dv_score = DataValidation(type="whole", operator="between", formula1="0", formula2="710", allow_blank=True)
    for dv in (dv_status, dv_listen, dv_read, dv_score):
        ws.add_data_validation(dv)
    dv_status.add("O5:O18")
    dv_listen.add("J5:J18")
    dv_read.add("K5:K18")
    dv_score.add("L5:L18")
    ws.conditional_formatting.add("I5:I18", CellIsRule(operator="lessThan", formula=["0.7"], fill=PatternFill("solid", fgColor=PALE_RED)))
    ws.conditional_formatting.add("I5:I18", CellIsRule(operator="greaterThanOrEqual", formula=["0.85"], fill=PatternFill("solid", fgColor=PALE_GREEN)))
    widths = [8, 11, 11, 30, 13, 13, 12, 12, 12, 14, 14, 13, 32, 38, 14]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "E5"
    ws.print_area = "A1:O18"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_mock_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("模考记录")
    ws.sheet_view.showGridLines = False
    style_title(ws, "模考记录", "原始正确数与出版社估分都只看趋势。CET 采用常模转换，不做“每题固定多少分”的简单换算。", 14)
    headers = ["日期", "试卷", "性质", "写作自评/15", "听力正确/25", "阅读正确/30", "翻译自评/15", "出版社估分", "用时(min)", "是否中断", "最大失分点", "48h复测日期", "阶段目标", "备注"]
    table_header(ws, 4, 1, headers)
    mock_rows = [
        (date(2026, 9, 12), "F0", "基线", "记录真实基线"),
        (date(2026, 10, 31), "M01", "全真", "≥380"),
        (date(2026, 11, 7), "M02", "全真", "≈400"),
        (date(2026, 11, 15), "M03", "全真", "≥415"),
        (date(2026, 11, 21), "M04", "全真", "≥425"),
        (date(2026, 11, 28), "M05", "全真", "≥435"),
        (date(2026, 12, 5), "M06", "全真", "≥450，或近两套均≥425"),
    ]
    for r_idx, (d, code, nature, target) in enumerate(mock_rows, 5):
        values = [d, code, nature, None, None, None, None, None, None, "否", None, d + timedelta(days=2), target, None]
        for c_idx, value in enumerate(values, 1):
            c = ws.cell(r_idx, c_idx, value)
            c.font = Font(name="Microsoft YaHei", size=9, bold=(c_idx in (1, 2, 13)), color=INK)
            c.fill = PatternFill("solid", fgColor=PALE_AMBER if code == "F0" else PALE_GREEN)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center" if c_idx <= 10 or c_idx in (12, 13) else "left")
            if isinstance(value, date):
                c.number_format = "yyyy-mm-dd"
        ws.row_dimensions[r_idx].height = 46
    end_row = 4 + len(mock_rows)
    validations = [
        (DataValidation(type="whole", operator="between", formula1="0", formula2="15", allow_blank=True), f"D5:D{end_row}"),
        (DataValidation(type="whole", operator="between", formula1="0", formula2="25", allow_blank=True), f"E5:E{end_row}"),
        (DataValidation(type="whole", operator="between", formula1="0", formula2="30", allow_blank=True), f"F5:F{end_row}"),
        (DataValidation(type="whole", operator="between", formula1="0", formula2="15", allow_blank=True), f"G5:G{end_row}"),
        (DataValidation(type="whole", operator="between", formula1="0", formula2="710", allow_blank=True), f"H5:H{end_row}"),
        (DataValidation(type="whole", operator="between", formula1="0", formula2="240", allow_blank=True), f"I5:I{end_row}"),
        (DataValidation(type="list", formula1='"否,是"', allow_blank=False), f"J5:J{end_row}"),
    ]
    for dv, rng in validations:
        ws.add_data_validation(dv)
        dv.add(rng)
    ws.conditional_formatting.add(f"H5:H{end_row}", CellIsRule(operator="greaterThanOrEqual", formula=["425"], fill=PatternFill("solid", fgColor=PALE_GREEN)))
    ws.conditional_formatting.add(f"H5:H{end_row}", CellIsRule(operator="lessThan", formula=["400"], fill=PatternFill("solid", fgColor=PALE_RED)))
    widths = [13, 10, 10, 15, 15, 15, 15, 15, 12, 12, 34, 15, 27, 30]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "D5"
    ws.print_area = f"A1:N{end_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_error_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("错题记录")
    ws.sheet_view.showGridLines = False
    style_title(ws, "错题记录", "只记会重复犯、值得复测的题。每条必须写“根因 + 下次动作”，不要抄长解析。", 14)
    headers = ["编号", "日期", "来源", "模块", "题型", "我的答案", "正确答案", "根因", "证据/关键句", "下次动作", "48h复测", "7天复测", "掌握状态", "备注"]
    table_header(ws, 4, 1, headers)
    for r in range(5, 105):
        ws.cell(r, 1, f'=IF(B{r}="","",ROW()-4)')
        for c in range(1, 15):
            cell = ws.cell(r, c)
            cell.font = Font(name="Microsoft YaHei", size=9, color=INK)
            cell.fill = PatternFill("solid", fgColor=WHITE if r % 2 else OFFWHITE)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 42
    ws.cell(5, 2).comment = Comment("示例：2026-09-12。删除本提示不影响表格。", "Codex")
    ws.cell(5, 8).comment = Comment("根因要可行动，例如“没听到转折词”，不要只写“粗心”。", "Codex")
    ws.cell(5, 10).comment = Comment("示例：下次先圈定位词；听到 however 立即回到题干。", "Codex")
    dvs = [
        (DataValidation(type="list", formula1='"听力,阅读,写作,翻译,词汇"', allow_blank=True), "D5:D104"),
        (DataValidation(type="list", formula1='"长对话,篇章听力,讲座讲话,选词填空,长篇匹配,仔细阅读,作文,翻译,其他"', allow_blank=True), "E5:E104"),
        (DataValidation(type="list", formula1='"词汇,听音辨识,走神,定位,同义替换,长难句,逻辑,速度,审题,语法,搭配,表达,其他"', allow_blank=True), "H5:H104"),
        (DataValidation(type="list", formula1='"未复习,48h正确,7天正确,已掌握,仍需复测"', allow_blank=True), "M5:M104"),
    ]
    for dv, rng in dvs:
        ws.add_data_validation(dv)
        dv.add(rng)
    ws.conditional_formatting.add("M5:M104", FormulaRule(formula=['$M5="已掌握"'], fill=PatternFill("solid", fgColor=PALE_GREEN)))
    ws.conditional_formatting.add("M5:M104", FormulaRule(formula=['OR($M5="未复习",$M5="仍需复测")'], fill=PatternFill("solid", fgColor=PALE_RED)))
    widths = [8, 13, 13, 11, 15, 17, 17, 18, 42, 38, 14, 14, 14, 30]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "F5"
    ws.auto_filter.ref = "A4:N104"
    ws.print_title_rows = "1:4"
    ws.print_area = "A1:N104"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_instructions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("使用说明")
    ws.sheet_view.showGridLines = False
    style_title(ws, "怎么用这份计划", "建议先看 5 分钟，再从“每日计划”开始。所有浅黄色单元格都表示需要确认或弹性处理。", 8)
    sections = [
        (4, "每天", [
            "打开“每日计划”，只看当天一行；完成后把状态改为“已完成”，填写实际分钟和 1–5 分自评。",
            "任务中的分钟数是时间上限。到点停止比为了做完题拖到深夜更重要。",
            "当天漏做：不整体顺延。只带走最重要的一项，放到周日复盘；最多 1 项。",
        ]),
        (10, "每周", [
            "周日填写“每周复盘”：完成率、听力/阅读原始正确数、最大问题、下周唯一调整。",
            "完成率低于 70%：下周总量降 15%，但保留听力和仔细阅读；高于 85% 才考虑加量。",
            "模考后的 48 小时内完成错题二测；7 天后只复测仍然不稳的题。",
        ]),
        (16, "真题", [
            "购买后先去“真题映射”填实际年份/套次。F0 是基线，D01–D05 可拆，M01–M06 为保留模考卷。",
            "全真模拟 130 分钟：写作 30、听力约 30、阅读 40、翻译 30；不中断、不查词。考试具体流程以当次考生须知为准。",
            "出版社估分只看趋势。真正的考前信号是：近两套都 425+，且听力/阅读原始正确率稳定。",
        ]),
        (22, "忙碌/生病", [
            "最低可执行版 40 分钟：到期词 10 + 错句盲听 15 + 仔细阅读证据定位 15。",
            "旅行 10/2–10/5、运动会减量任务均不补课；恢复日从当日计划直接继续。",
            "若连续三天无法完成 70%，先砍掉选词填空和额外写译，不砍真题听力与仔细阅读。",
        ]),
        (28, "考试日期变更", [
            "当前 12/12 是倒排占位，不是教育部教育考试院 2026 下半年正式公告。",
            "若正式日期晚于 12/12：12/12 做轻量全真或 M06 复测，然后重复第 14 周；若提前，从考前 6 天直接执行第 14 周。",
            "9 月起关注莆田学院报名通知；正式公告发布后，修改总览日期和每日计划末段即可。",
        ]),
    ]
    for start_row, title, bullets in sections:
        section_header(ws, start_row, title, 1, 8, TEAL if title != "考试日期变更" else AMBER)
        for i, bullet in enumerate(bullets, start_row + 1):
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
            c = ws.cell(i, 1, "• " + bullet)
            c.font = Font(name="Microsoft YaHei", size=10, color=INK)
            c.fill = PatternFill("solid", fgColor=WHITE if i % 2 else OFFWHITE)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center", wrap_text=True)
            ws.row_dimensions[i].height = 34

    section_header(ws, 34, "来源与核实日期", 1, 8, NAVY)
    sources = [
        ("教育部教育考试院 CET 考试动态", NEEA_NEWS_URL, "截至 2026-08-27 未见 2026 下半年正式报名公告"),
        ("CET 报名网考试时间", NEEA_TEST_URL, "截至 2026-08-27 仍只列 2026 上半年安排"),
        ("CET 笔试介绍", NEEA_CET_URL, "官方说明笔试每年 6 月、12 月举行"),
        ("2026 上半年官方公告", NEEA_2026H1_URL, "用于参考近次 CET-6 时段 15:00–17:25；下半年仍待确认"),
        ("国务院 2026 年节假日通知", HOLIDAY_URL, "中秋 9/25–27；国庆 10/1–7，9/20、10/10 调休上班"),
        ("上海市年度考试计划", SHANGHAI_PLAN_URL, "提供 12/12 排程依据，但不是本次中央正式公告，可能调整"),
        ("莆田学院 2026–2027 第一学期校历", "用户提供图片", "9/7 开学；9/25 中秋；11/12–14 运动会"),
    ]
    table_header(ws, 35, 1, ["来源", "链接/出处", "本计划如何使用"], fill=NAVY)
    for r_idx, row in enumerate(sources, 36):
        for c_idx, value in enumerate(row, 1):
            c = ws.cell(r_idx, c_idx, value)
            c.font = Font(name="Microsoft YaHei", size=9, color=INK)
            c.fill = PatternFill("solid", fgColor=PALE_AMBER if "年度考试计划" in row[0] else WHITE)
            c.border = thin_border()
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if c_idx == 2 and isinstance(value, str) and value.startswith("http"):
                c.hyperlink = value
                c.style = "Hyperlink"
        ws.row_dimensions[r_idx].height = 42
    for col, width in {"A":24, "B":52, "C":68, "D":14, "E":14, "F":14, "G":14, "H":14}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A4"
    ws.print_area = "A1:H42"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 2
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_excel(plans: List[DayPlan]) -> None:
    wb = Workbook()
    wb.properties.creator = "Codex"
    wb.properties.title = "2026年12月CET-6备考计划（初版）"
    wb.properties.subject = "莆田学院学生 CET-6 425+ 备考"
    wb.properties.description = "2026-09-07 起，按 450 能力倒排；12-12 为待官方确认占位日。"
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass
    build_overview(wb, plans)
    build_daily_sheet(wb, plans)
    build_resources_sheet(wb)
    build_paper_map_sheet(wb)
    build_weekly_review_sheet(wb)
    build_mock_sheet(wb)
    build_error_sheet(wb)
    build_instructions_sheet(wb)
    for ws in wb.worksheets:
        set_all_font(ws)
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.oddFooter.left.text = "2026 CET-6 425+ 计划"
        ws.oddFooter.left.size = 8
        ws.oddFooter.left.color = MUTED
    wb.save(XLSX_PATH)


def register_pdf_fonts() -> Tuple[str, str]:
    regular = Path(r"C:\Windows\Fonts\msyh.ttc")
    bold = Path(r"C:\Windows\Fonts\msyhbd.ttc")
    if not regular.exists():
        raise FileNotFoundError(regular)
    pdfmetrics.registerFont(TTFont("MSYH", str(regular), subfontIndex=0))
    pdfmetrics.registerFont(TTFont("MSYH-Bold", str(bold if bold.exists() else regular), subfontIndex=0))
    pdfmetrics.registerFontFamily("MSYH", normal="MSYH", bold="MSYH-Bold", italic="MSYH", boldItalic="MSYH-Bold")
    return "MSYH", "MSYH-Bold"


class NumberedDocTemplate(BaseDocTemplate):
    def __init__(self, filename, **kwargs):
        super().__init__(filename, **kwargs)
        frame = Frame(self.leftMargin, self.bottomMargin, self.width, self.height, id="normal")
        self.addPageTemplates([PageTemplate(id="all", frames=frame, onPage=self._header_footer)])

    def _header_footer(self, canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#D5DEE5"))
        canvas.setLineWidth(0.5)
        canvas.line(self.leftMargin, A4[1] - 14 * mm, A4[0] - self.rightMargin, A4[1] - 14 * mm)
        canvas.setFont("MSYH", 7.5)
        canvas.setFillColor(colors.HexColor("#667085"))
        canvas.drawString(self.leftMargin, A4[1] - 10.5 * mm, "2026 年 12 月 CET-6 425+ 备考计划 · 初版 v1.0")
        canvas.drawRightString(A4[0] - self.rightMargin, 9 * mm, f"第 {doc.page} 页")
        canvas.restoreState()


def make_pdf_styles():
    styles = getSampleStyleSheet()
    return {
        "cover_title": ParagraphStyle("CoverTitle", fontName="MSYH-Bold", fontSize=24, leading=31, textColor=colors.HexColor("#173B57"), alignment=TA_CENTER, spaceAfter=8 * mm),
        "cover_sub": ParagraphStyle("CoverSub", fontName="MSYH", fontSize=12, leading=18, textColor=colors.HexColor("#667085"), alignment=TA_CENTER),
        "h1": ParagraphStyle("H1", fontName="MSYH-Bold", fontSize=18, leading=24, textColor=colors.HexColor("#173B57"), spaceAfter=5 * mm),
        "h2": ParagraphStyle("H2", fontName="MSYH-Bold", fontSize=12, leading=16, textColor=colors.HexColor("#145B68"), spaceBefore=3 * mm, spaceAfter=2 * mm),
        "body": ParagraphStyle("Body", fontName="MSYH", fontSize=9.2, leading=14, textColor=colors.HexColor("#243447"), spaceAfter=2 * mm),
        "small": ParagraphStyle("Small", fontName="MSYH", fontSize=7.8, leading=10.8, textColor=colors.HexColor("#243447")),
        "small_muted": ParagraphStyle("SmallMuted", fontName="MSYH", fontSize=7.5, leading=10.5, textColor=colors.HexColor("#667085")),
        "table_head": ParagraphStyle("TableHead", fontName="MSYH-Bold", fontSize=8, leading=10, textColor=colors.white, alignment=TA_CENTER),
        "table_cell": ParagraphStyle("TableCell", fontName="MSYH", fontSize=7.4, leading=10, textColor=colors.HexColor("#243447")),
        "table_cell_center": ParagraphStyle("TableCellCenter", fontName="MSYH", fontSize=7.5, leading=10, textColor=colors.HexColor("#243447"), alignment=TA_CENTER),
        "note": ParagraphStyle("Note", fontName="MSYH", fontSize=8.5, leading=13, textColor=colors.HexColor("#7A4D00"), backColor=colors.HexColor("#FFF3D8"), borderColor=colors.HexColor("#D99B3D"), borderWidth=0.6, borderPadding=7, spaceAfter=4 * mm),
    }


def p(text: str, style) -> Paragraph:
    safe = (text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br/>")
    return Paragraph(safe, style)


def pdf_table(data, col_widths, style_commands, repeat_rows=1, row_heights=None):
    t = PDFTable(data, colWidths=col_widths, rowHeights=row_heights, repeatRows=repeat_rows, hAlign="LEFT")
    base = [
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D5DEE5")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    t.setStyle(TableStyle(base + style_commands))
    return t


def pdf_week_page(plans: List[DayPlan], week_no: int, styles) -> List:
    week_plans = [x for x in plans if x.week_no == week_no]
    theme, goal, paper = WEEK_SPECS[week_no]
    start, end = week_plans[0].day, week_plans[-1].day
    elements = [
        Paragraph(f"第 {week_no} 周｜{theme}", styles["h1"]),
        Paragraph(f"{start.strftime('%m/%d')}–{end.strftime('%m/%d')} · 主线材料：{paper} · 本周验收：{goal}", styles["note"]),
    ]
    data = [[p("日期", styles["table_head"]), p("时长", styles["table_head"]), p("日型", styles["table_head"]), p("具体任务", styles["table_head"]), p("完成标准", styles["table_head"])]]
    for item in week_plans:
        task_parts = []
        for label, value in [("词", item.word), ("听", item.listening), ("读", item.reading), ("写译", item.writing_translation), ("复盘", item.review)]:
            if value:
                task_parts.append(f"<b>{label}</b> {value}")
        task_html = "<br/>".join(task_parts) if task_parts else "—"
        row = [
            Paragraph(f"<b>{item.day.strftime('%m/%d')}</b><br/>{WEEKDAY_CN[item.day.weekday()]}", styles["table_cell_center"]),
            Paragraph(f"<b>{item.planned_minutes}</b><br/>min" if item.planned_minutes else "0", styles["table_cell_center"]),
            p(item.day_type, styles["table_cell_center"]),
            Paragraph(task_html, styles["table_cell"]),
            p(item.success, styles["table_cell"]),
        ]
        data.append(row)
    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#173B57")),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
    ]
    for idx, item in enumerate(week_plans, 1):
        bg = colors.white
        if item.day_type in ("旅行·完全留空",):
            bg = colors.HexColor("#E9EDF1")
        elif "轻量" in item.day_type or "仅词汇" in item.day_type:
            bg = colors.HexColor("#FFF3D8")
        elif item.day_type == "全真模考":
            bg = colors.HexColor("#E6F3EC")
        elif item.day_type == "模考复盘":
            bg = colors.HexColor("#EDF6F8")
        elif item.day_type == "暂定考试日":
            bg = colors.HexColor("#FBE9E9")
        elif item.day.weekday() >= 5:
            bg = colors.HexColor("#F7F9FB")
        commands.append(("BACKGROUND", (0, idx), (-1, idx), bg))
    elements.append(pdf_table(data, [20 * mm, 13 * mm, 25 * mm, 91 * mm, 40 * mm], commands))
    elements.append(Spacer(1, 3 * mm))
    elements.append(Paragraph("打卡：□ 本周完成率 ≥ 70%　□ 至少一次听力错句二测　□ 至少一次阅读证据链复盘　□ 下周只调整一件事", styles["small_muted"]))
    return elements


def build_pdf(plans: List[DayPlan]) -> None:
    register_pdf_fonts()
    styles = make_pdf_styles()
    doc = NumberedDocTemplate(str(PDF_PATH), pagesize=A4, rightMargin=12 * mm, leftMargin=12 * mm, topMargin=20 * mm, bottomMargin=16 * mm, title="2026年12月CET-6备考计划", author="Codex")
    story: List = []

    # Cover.
    story.extend([
        Spacer(1, 38 * mm),
        Paragraph("2026 年 12 月<br/>CET-6 425+ 备考计划", styles["cover_title"]),
        Paragraph("按 450 分能力构建容错｜Excel + PDF 初版", styles["cover_sub"]),
        Spacer(1, 16 * mm),
    ])
    score_data = [
        [p("模块", styles["table_head"]), p("当前", styles["table_head"]), p("训练目标", styles["table_head"]), p("策略", styles["table_head"])],
        [p("听力", styles["table_cell_center"]), p("78", styles["table_cell_center"]), p("150", styles["table_cell_center"]), p("真题原音；讲座/讲话与错句精听优先", styles["table_cell"])],
        [p("阅读", styles["table_cell_center"]), p("86", styles["table_cell_center"]), p("165", styles["table_cell_center"]), p("仔细阅读 + 长篇匹配优先", styles["table_cell"])],
        [p("写作+翻译", styles["table_cell_center"]), p("117", styles["table_cell_center"]), p("135", styles["table_cell_center"]), p("每周完整输出一次，维持提分", styles["table_cell"])],
        [p("总分", styles["table_cell_center"]), p("281", styles["table_cell_center"]), p("450", styles["table_cell_center"]), p("目标至少 425；以 450 训练留容错", styles["table_cell"])],
    ]
    story.append(pdf_table(score_data, [32 * mm, 24 * mm, 30 * mm, 92 * mm], [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#173B57")),
        ("BACKGROUND", (0, 4), (-1, 4), colors.HexColor("#E6F3EC")),
    ]))
    story.extend([
        Spacer(1, 12 * mm),
        Paragraph("适用：莆田学院 · 开始：2026-09-07 · 工作日 90–120 分钟 · 周末 150–180 分钟", styles["cover_sub"]),
        Spacer(1, 9 * mm),
        Paragraph("重要：2026 年下半年官方笔试日期截至 2026-08-27 尚未发布。本计划暂以 12 月 12 日作为倒排占位，最终以教育部教育考试院、学校通知和准考证为准。", styles["note"]),
        Spacer(1, 22 * mm),
        Paragraph("版本 v1.0 · 生成于 2026-08-27", styles["small_muted"]),
        PageBreak(),
    ])

    # Quick start.
    story.extend([
        Paragraph("先看这一页：怎么执行", styles["h1"]),
        Paragraph("每天只做当天一行。任务里的分钟数是上限；到点停止，避免计划把正常课程和睡眠吞掉。", styles["body"]),
        Paragraph("三条硬规则", styles["h2"]),
    ])
    rules_data = [
        [p("1", styles["table_head"]), p("不滚雪球", styles["table_head"]), p("漏做不整体顺延；最多带 1 个核心任务到周日。", styles["table_head"])],
        [p("2", styles["table_cell_center"]), p("最低版 40 分钟", styles["table_cell"]), p("到期词 10 + 错句盲听 15 + 仔细阅读证据定位 15。", styles["table_cell"])],
        [p("3", styles["table_cell_center"]), p("模考不偷看", styles["table_cell"]), p("M01–M06 在指定日期前不看答案、不听原文；完整模拟 130 分钟。", styles["table_cell"])],
    ]
    story.append(pdf_table(rules_data, [12 * mm, 40 * mm, 126 * mm], [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#145B68")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F7F9FB")),
    ], repeat_rows=0))
    story.extend([
        Paragraph("必须保留的特殊日期", styles["h2"]),
    ])
    special_data = [
        [p("日期", styles["table_head"]), p("安排", styles["table_head"]), p("理由", styles["table_head"])],
        [p("9/25", styles["table_cell_center"]), p("35 分钟轻量", styles["table_cell"]), p("中秋节当天只做词汇与已学错句。", styles["table_cell"])],
        [p("9/30–10/1", styles["table_cell_center"]), p("仅词汇 15–20 分钟", styles["table_cell"]), p("新加坡旅行缓冲；紧张可跳过且不补。", styles["table_cell"])],
        [p("10/2–10/5", styles["table_cell_center"]), p("完全留空", styles["table_cell"]), p("任何内容都不安排，也不形成欠账。", styles["table_cell"])],
        [p("11/12–11/14", styles["table_cell_center"]), p("15–20 分钟词汇", styles["table_cell"]), p("校历明确为运动会，只维持词感。", styles["table_cell"])],
        [p("12/12", styles["table_cell_center"]), p("暂定考试日", styles["table_cell"]), p("排程占位；待教育部教育考试院正式确认。", styles["table_cell"])],
    ]
    story.append(pdf_table(special_data, [33 * mm, 50 * mm, 95 * mm], [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#173B57")),
        ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#E9EDF1")),
        ("BACKGROUND", (0, 5), (-1, 5), colors.HexColor("#FBE9E9")),
    ]))
    story.extend([
        Paragraph("考前达标信号", styles["h2"]),
        Paragraph("最近两套出版社估分均达到 425+，且至少一套接近 450；听力稳定在约 17/25、阅读约 21/30（仅作训练方向，不等于固定分值换算）；作文和翻译均能在 30 分钟内完成。", styles["body"]),
        PageBreak(),
    ])

    # Resources.
    story.extend([
        Paragraph("资料组合：性价比优先", styles["h1"]),
        Paragraph("纸质只需要一套真题。真正拉开差距的是同一套题做完、订正、48 小时复测，而不是买更多书。", styles["body"]),
    ])
    resource_data = [
        [p("资料", styles["table_head"]), p("必要性", styles["table_head"]), p("具体用法", styles["table_head"]), p("费用", styles["table_head"])],
        [p("最新版六级真题详解（星火优先；新东方同类版可替代）", styles["table_cell"]), p("必备", styles["table_cell_center"]), p("确认 ≥12 套完整真题、原版音频、听力原文和逐题解析；按 F0/D/M 编号。不要买两套。", styles["table_cell"]), p("约 35–60 元", styles["table_cell_center"])],
        [p("不背单词", styles["table_cell"]), p("必备", styles["table_cell_center"]), p("前期新词 35–40 + 到期复习；后期只复习真题错词。墨墨/百词斩可替代，但只选一个。", styles["table_cell"]), p("基础功能可免费", styles["table_cell_center"])],
        [p("随书音频 + 每日英语听力", styles["table_cell"]), p("推荐", styles["table_cell_center"]), p("真题原音优先；用逐句回放、0.9×/1.0×、文本对照。只听写 6–8 个关键句，不做整篇逐字听写。", styles["table_cell"]), p("随书/基础免费", styles["table_cell_center"])],
        [p("本 Excel 错题/模考记录", styles["table_cell"]), p("必备", styles["table_cell_center"]), p("每题只记根因、证据、下次动作、48 小时和 7 天复测。", styles["table_cell"]), p("免费", styles["table_cell_center"])],
        [p("现有 AI/学校批改网", styles["table_cell"]), p("推荐", styles["table_cell_center"]), p("先独立限时写，再按 CET-6 维度纠错；要求给理由，不让它直接代写。", styles["table_cell"]), p("已有服务/校内", styles["table_cell_center"])],
        [p("BBC 6 Minute English", styles["table_cell"]), p("可选", styles["table_cell_center"]), p("旅行或轻量日补充，每周最多一次；真题没完成时不听。", styles["table_cell"]), p("免费", styles["table_cell_center"])],
    ]
    story.append(pdf_table(resource_data, [44 * mm, 19 * mm, 91 * mm, 24 * mm], [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#173B57")),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#E6F3EC")),
        ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#EDF6F8")),
    ]))
    story.extend([
        Spacer(1, 4 * mm),
        Paragraph("推荐购买结论：只买 1 套最新版六级真题（约 35–60 元）。如果已经能坚持用 App，就不必再买词汇书；也不建议额外购买阅读、听力、作文三本专项书。", styles["note"]),
        Paragraph("电子资料不只背词：随书音频负责真题原音，每日英语听力负责逐句回放，现有 AI/批改网负责写译纠错，本 Excel 负责错题与模考趋势，官方 CET 网站负责日期和报名。", styles["body"]),
        PageBreak(),
    ])

    # Phase map.
    story.extend([
        Paragraph("14 周路线与里程碑", styles["h1"]),
    ])
    phase_data = [[p("周", styles["table_head"]), p("主题", styles["table_head"]), p("验收重点", styles["table_head"]), p("材料", styles["table_head"])]]
    for week_no in range(1, 15):
        theme, goal, paper = WEEK_SPECS[week_no]
        phase_data.append([p(str(week_no), styles["table_cell_center"]), p(theme, styles["table_cell"]), p(goal, styles["table_cell"]), p(paper, styles["table_cell_center"])])
    commands = [("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#173B57"))]
    for i in range(1, 15):
        commands.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F7F9FB" if i % 2 == 0 else "#FFFFFF")))
    story.append(pdf_table(phase_data, [12 * mm, 52 * mm, 98 * mm, 16 * mm], commands))
    story.extend([
        Spacer(1, 4 * mm),
        Paragraph("分数目标是导航，不是惩罚。若某周低于目标，只选择一个最大短板修复；不要同时增加词汇、听力、阅读和写译总量。", styles["note"]),
        PageBreak(),
    ])

    # Weekly daily pages.
    for week_no in range(1, 15):
        story.extend(pdf_week_page(plans, week_no, styles))
        if week_no != 14:
            story.append(PageBreak())

    story.append(PageBreak())
    story.extend([
        Paragraph("复盘、估分与错题闭环", styles["h1"]),
        Paragraph("每次模考后，不要立刻找下一套。分数上涨来自把同一种错误关掉。", styles["body"]),
        Paragraph("模考节点", styles["h2"]),
    ])
    mock_data = [[p("试卷", styles["table_head"]), p("日期", styles["table_head"]), p("阶段目标", styles["table_head"]), p("次日动作", styles["table_head"])]]
    mock_entries = [
        ("F0", "9/12", "真实基线", "四模块分类错因"),
        ("M01", "10/31", "≥380", "听读错题二测"),
        ("M02", "11/7", "≈400", "只修最大两项"),
        ("M03", "11/15", "≥415", "11/16 深度复盘"),
        ("M04", "11/21", "≥425", "确认是否首次越线"),
        ("M05", "11/28", "≥435", "检查是否连续稳定"),
        ("M06", "12/5", "≥450 或近两套均≥425", "最后一次完整复盘"),
    ]
    for row in mock_entries:
        mock_data.append([p(x, styles["table_cell_center"] if i < 3 else styles["table_cell"]) for i, x in enumerate(row)])
    story.append(pdf_table(mock_data, [20 * mm, 25 * mm, 55 * mm, 78 * mm], [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#173B57")),
        ("BACKGROUND", (0, 5), (-1, 7), colors.HexColor("#E6F3EC")),
    ]))
    story.extend([
        Paragraph("错题五步", styles["h2"]),
        Paragraph("1）写模块与题型；2）选一个真实根因；3）抄最短证据/关键句；4）写下一次动作；5）48 小时与 7 天复测。能连续两次独立做对，才标“已掌握”。", styles["body"]),
        Paragraph("不要做的事", styles["h2"]),
        Paragraph("不要把整篇解析抄进错题本；不要用新的模拟题替代真题复盘；不要用原始正确数直接乘固定分值估算 710 分制；不要因一次低分临时换资料。", styles["note"]),
        PageBreak(),
    ])

    # Sources and official-date note.
    story.extend([
        Paragraph("日期依据与待确认事项", styles["h1"]),
        Paragraph("截至 2026-08-27，教育部教育考试院 CET 动态页和报名网考试时间页均未发布 2026 年下半年正式公告。12/12 仅作倒排占位；上海市年度考试计划提供了参考，但该计划不是中央当次正式公告，且历史上年度计划可能调整。", styles["note"]),
    ])
    source_rows = [
        ("教育部教育考试院 CET 动态", NEEA_NEWS_URL, "核对下半年正式公告"),
        ("CET 报名网考试时间", NEEA_TEST_URL, "截至核实时仅列上半年"),
        ("CET 笔试介绍", NEEA_CET_URL, "官方说明每年 6 月、12 月举行"),
        ("2026 上半年官方公告", NEEA_2026H1_URL, "只用于参考近次六级时段"),
        ("国务院 2026 节假日通知", HOLIDAY_URL, "中秋、国庆和补班安排"),
        ("上海市年度考试计划", SHANGHAI_PLAN_URL, "12/12 排程参考，非最终官宣"),
        ("莆田学院第一学期校历", "用户提供图片", "9/7 开学、9/25 中秋、11/12–14 运动会"),
    ]
    src_data = [[p("来源", styles["table_head"]), p("地址/出处", styles["table_head"]), p("用途", styles["table_head"])]]
    for name, url, use in source_rows:
        link_text = url if not url.startswith("http") else f'<link href="{url}" color="#1F7A8C">{url}</link>'
        src_data.append([p(name, styles["table_cell"]), Paragraph(link_text, styles["small_muted"]), p(use, styles["table_cell"])])
    story.append(pdf_table(src_data, [42 * mm, 96 * mm, 40 * mm], [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#173B57")),
        ("BACKGROUND", (0, 6), (-1, 6), colors.HexColor("#FFF3D8")),
    ]))
    story.extend([
        Spacer(1, 5 * mm),
        Paragraph("发布归类：个人学习资料。生成源码、Excel 与 PDF 归档于 GitHub 仓库 adaml000612-web/personal-web-lab 的 projects/cet6-study-plan-2026/；原始成绩与校历截图因含个人信息未上传。", styles["small_muted"]),
    ])

    doc.build(story)


def validate_outputs(plans: List[DayPlan]) -> None:
    assert len(plans) == (PLACEHOLDER_EXAM_DATE - START_DATE).days + 1 == 97
    protected = [p for p in plans if date(2026, 10, 2) <= p.day <= date(2026, 10, 5)]
    assert all(p.planned_minutes == 0 and not any([p.word, p.listening, p.reading, p.writing_translation]) for p in protected)
    sports = [p for p in plans if date(2026, 11, 12) <= p.day <= date(2026, 11, 14)]
    assert all(p.planned_minutes <= 20 and p.word and not p.listening and not p.reading and not p.writing_translation for p in sports)
    assert XLSX_PATH.exists() and XLSX_PATH.stat().st_size > 50_000
    assert PDF_PATH.exists() and PDF_PATH.stat().st_size > 50_000

    wb = load_workbook(XLSX_PATH, data_only=False, read_only=False)
    expected = ["总览", "每日计划", "资料清单", "真题映射", "每周复盘", "模考记录", "错题记录", "使用说明"]
    assert wb.sheetnames == expected, wb.sheetnames
    ws = wb["每日计划"]
    assert ws.max_row == 98
    date_to_row = {ws.cell(r, 1).value.date() if isinstance(ws.cell(r, 1).value, datetime) else ws.cell(r, 1).value: r for r in range(2, ws.max_row + 1)}
    for d in [date(2026, 10, 2), date(2026, 10, 3), date(2026, 10, 4), date(2026, 10, 5)]:
        r = date_to_row[d]
        assert ws.cell(r, 6).value == 0
        assert ws.cell(r, 14).value == "计划休息"
    assert ws.cell(date_to_row[date(2026, 11, 12)], 5).value == "运动会·轻量"
    assert ws.cell(date_to_row[PLACEHOLDER_EXAM_DATE], 14).value == "待确认"
    wb.close()

    from pypdf import PdfReader
    reader = PdfReader(str(PDF_PATH))
    assert len(reader.pages) >= 19
    assert reader.metadata is not None


def main() -> None:
    plans = build_daily_plans()
    build_excel(plans)
    build_pdf(plans)
    validate_outputs(plans)
    planned_minutes = sum(p.planned_minutes for p in plans)
    print(f"Created: {XLSX_PATH}")
    print(f"Created: {PDF_PATH}")
    print(f"Days: {len(plans)}; planned hours: {planned_minutes / 60:.1f}")


if __name__ == "__main__":
    main()
